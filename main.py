import os

from selenium import webdriver
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook


def get_website():
    driver.maximize_window()
    driver.get("https://www.thsrc.com.tw/")


def get_condition(driver):
    # XPath
    cookie_agree_xpath = '/html/body/div[6]/div/div[3]/button[2]'
    start_stop_id = "select_location01"
    end_stop_id = "select_location02"
    one_two_way_id = "typesofticket"
    depart_date_id = "Departdate01"
    start_search_id = "start-search"

    # Agree cookie
    cookie_agree_button = driver.find_element_by_xpath(cookie_agree_xpath)
    cookie_agree_button.click()

    # Select station
    start_station = driver.find_element_by_id(start_stop_id)
    Select(start_station).select_by_visible_text("台北")

    end_station = driver.find_element_by_id(end_stop_id)
    Select(end_station).select_by_visible_text("桃園")

    # Select type of ticket
    one_two_way = driver.find_element_by_id(one_two_way_id)
    Select(one_two_way).select_by_visible_text("單程")

    # Select time
    depart_date_box = driver.find_element_by_id(depart_date_id)
    depart_date_box.click()
    time_table = driver.find_element_by_class_name("table-condensed")
    # In this date picker, a month is a 6*7 table. "tr" means rows, "rd" means columns.
    time_table.find_element_by_xpath('//*[@id="tot-1"]/div[1]/div/ul/li[1]/div/div[1]/table/tbody/tr[6]/td[1]').click()

    # Start search
    start_search_button = driver.find_element_by_id(start_search_id)
    start_search_button.click()


def get_content(driver):
    def export_to_excel(p_time):
        # Generate excel file
        wb = Workbook()
        ws = wb.create_sheet("Mysheet", 0)
        i = 3

        ws['A1'] = "台北"
        ws['B1'] = "到"
        ws['C1'] = "桃園"
        ws['A2'] = "出發時間"
        ws['C2'] = "抵達時間"
        for time in p_time:
            ws[f'A{i}'] = time[0]
            ws[f'B{i}'] = '-->'
            ws[f'C{i}'] = time[1]
            i += 1

        os.makedirs("output", exist_ok=True)
        wb.save('output/Schedule.xlsx')

    # Set implicit wait for element finding
    driver.implicitly_wait(5)

    # Get content
    time_content = driver.find_elements_by_class_name("font-16r")
    time_table = [i.text for i in time_content]
    paired_time = [time_table[i:i + 2] for i in range(0, len(time_table), 2)]
    for pair in paired_time:
        print(f"{pair[0]} --> {pair[1]}")

    # Reset wait
    driver.implicitly_wait(0)

    # Export to Excel
    export_to_excel(paired_time)


def exit_browser(driver):
    driver.quit()


if __name__ == '__main__':
    # define browser driver
    chromedriver = "chromedriver/chromedriver.exe"
    driver = webdriver.Chrome(chromedriver)

    # Start working
    get_website()
    get_condition(driver)
    get_content(driver)

    # Stop working
    exit_browser(driver)

